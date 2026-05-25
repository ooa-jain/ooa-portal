from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file, flash
from pymongo import MongoClient
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from bson import ObjectId
import json, os, io, openpyxl, random
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'jain_ooa_semreadiness_2025_secret'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

MONGO_URI = "mongodb+srv://santoshks_db_user:viefoCaPp3CMCqTq@cluster0.v8wfkok.mongodb.net/formcraft?retryWrites=true&w=majority&appName=Cluster0"
client = MongoClient(MONGO_URI)
db = client['semreadiness']
submissions_col = db['submissions']
users_col = db['users']
settings_col = db['settings']

ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'admin2023'

DEPARTMENTS = [
    "Department of Computer Science and Engineering",
    "Department of Information Science and Engineering",
    "Department of Aerospace Engineering",
    "Department of Civil Engineering",
    "Department of Mechanical Engineering",
    "Department of Electrical and Electronics Engineering",
    "Department of Electronics and Communication Engineering",
    "Department of Food Technology",
    "Department of Humanities & Social Sciences",
    "Department of CERSSE",
    "Department of SSER",
    "Department of Jainology",
    "Department of Marine Science",
    "Department of Economics",
    "Department of Performing Arts and Cultural Studies",
    "Department of Languages",
    "Department of Journalism and Mass Communication",
    "Department of Law",
    "Department of Chemistry and Biochemistry",
    "Department of Biotechnology and Genetics",
    "Department of Microbiology and Botany",
    "Department of Data Analytics and Mathematical Science",
    "Department of Forensic Science",
    "Department of Physics and Electronics",
    "Department of Psychology and Allied Sciences",
    "Department of Allied Healthcare and Sciences",
    "Department of Computer Science and IT",
    "Department of Animation and Virtual Reality",
    "Department of Commerce",
    "Department of Management Studies",
    "Department of Design",
    "Department of Art and Design",
]

HOD_SECTIONS = [
    {"title": "Section 1: Curriculum & Course Matrix", "items": [
        {"id": "1", "text": "Previous semester's closing report insights incorporated"},
        {"id": "2", "text": "Course matrix verified (codes, credits, CBCS, LTPE) against BoS docs"},
        {"id": "3", "text": "Course matrix uploaded on ERP"},
        {"id": "4", "text": "Open Elective courses cross-checked with offering departments"},
        {"id": "5", "text": "Students given ≥ 1 week to choose electives after orientation"},
        {"id": "6", "text": "Kochi / Online / ODL / other campus stakeholders included in planning", "kochi": True},
    ]},
    {"title": "Section 2: Faculty Allocation & Workload", "items": [
        {"id": "7", "text": "All courses have faculty assigned"},
        {"id": "8", "text": "Workload is fair and transparent across the department"},
        {"id": "9", "text": "Faculty informed of their course allocations formally"},
        {"id": "10", "text": "Faculty workload documented and assessed"},
    ]},
    {"title": "Section 3: TLEP Audit", "items": [
        {"id": "11", "text": "All faculty have submitted TLEPs"},
        {"id": "12", "text": "HOD has reviewed / audited all TLEPs"},
        {"id": "13", "text": "Innovative pedagogies and assessments collated from TLEPs"},
        {"id": "14", "text": "Previous semester faculty feedback shared with faculty"},
    ]},
    {"title": "Section 4: Timetable & ERP", "items": [
        {"id": "15", "text": "Master timetable prepared"},
        {"id": "16", "text": "Master timetable uploaded on ERP"},
        {"id": "17", "text": "Course creation enabled on ERP"},
        {"id": "18", "text": "All session plans uploaded on ERP by faculty"},
    ]},
    {"title": "Section 5: LMS & Digital Readiness", "items": [
        {"id": "19", "text": "All faculty trained on LMS platform"},
        {"id": "20", "text": "Faculty actively using LMS — courses live"},
    ]},
    {"title": "Section 6: Students & Activities", "items": [
        {"id": "21", "text": "Co-curricular and extra-curricular activities planned and calendarised"},
        {"id": "22", "text": "Student Progression and Graduation report reviewed"},
        {"id": "23", "text": "Departmental Academic Achievement report prepared"},
    ]},
    {"title": "Section 7: Faculty Development & Research", "items": [
        {"id": "24", "text": "Faculty development / capacity building activities planned"},
        {"id": "25", "text": "Research activities (conferences, seminars) planned and calendarised"},
    ]},
    {"title": "Section 8: Infrastructure & Approvals", "items": [
        {"id": "26", "text": "Lab equipment, requirements and budget assessed"},
        {"id": "27", "text": "All necessary approvals obtained from University HO"},
    ]},
]

# ── Mail Helper ──────────────────────────────────────────

def send_otp_email(to_email, otp):
    try:
        host = "smtp.gmail.com"
        port = 587
        user = "info.loginpanel@gmail.com"
        password = "wedbfepklgtwtugf"
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'Your Security Passcode OTP'
        msg['From'] = user
        msg['To'] = to_email
        
        html_content = f"""
        <html>
          <body style="font-family: Arial, sans-serif; background-color: #f0f3fa; padding: 20px; text-align: center;">
            <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
              <h2 style="color: #0A2558; margin-bottom: 20px;">Semester Readiness Portal</h2>
              <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">You requested to reset your security passcode.</p>
              <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">Please use the following 4-digit OTP to proceed:</p>
              
              <div style="background: #f8f9fb; border: 2px dashed #0A2558; border-radius: 8px; padding: 15px; font-size: 32px; font-weight: bold; letter-spacing: 8px; color: #0A2558; margin-bottom: 30px;">
                {otp}
              </div>
              
              <p style="color: #8892aa; font-size: 12px;">If you did not request this, please ignore this email.</p>
            </div>
          </body>
        </html>
        """
        
        part = MIMEText(html_content, 'html')
        msg.attach(part)
        
        server = smtplib.SMTP(host, port)
        server.starttls()
        server.login(user, password)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"SMTP Error: {e}")
        return False

# ── Routes ──────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_email' not in session:
        return redirect(url_for('login'))
    user_doc = users_col.find_one({'email': session['user_email']})
    user = {
        'email': session['user_email'], 
        'name': session['user_name'],
        'department': user_doc.get('department', '') if user_doc else '',
        'timeout_pref': user_doc.get('timeout_pref', 15) if user_doc else 15,
        'lock_enabled': user_doc.get('lock_enabled', False) if user_doc else False,
        'first_time_login': user_doc.get('first_time_login', True) if user_doc else False
    }
    settings = settings_col.find_one({'_id': 'global'}) or {
        'readiness_enabled': True, 'readiness_deadline': '',
        'closure_enabled': True, 'closure_deadline': ''
    }
    return render_template('dashboard.html', user=user, settings=settings, departments=DEPARTMENTS)

@app.route('/readiness')
def readiness():
    if 'user_email' not in session:
        return redirect(url_for('login'))
    user = {'email': session['user_email'], 'name': session['user_name']}
    return render_template('form.html', hod_sections=HOD_SECTIONS, departments=DEPARTMENTS, user=user)

@app.route('/closure')
def closure():
    if 'user_email' not in session:
        return redirect(url_for('login'))
    user = {'email': session['user_email'], 'name': session['user_name']}
    return render_template('closure.html', user=user, departments=DEPARTMENTS)

@app.route('/login')
def login():
    if 'user_email' in session:
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/api/submit', methods=['POST'])
def submit():
    data = request.json
    sub_id = data.get('_id')
    
    if '_id' in data:
        del data['_id']
        
    data['timestamp'] = datetime.utcnow().isoformat()
    
    if sub_id:
        existing = submissions_col.find_one({'_id': ObjectId(sub_id)})
        if not existing:
            return jsonify({'ok': False, 'error': 'Submission not found'})
            
        edit_count = existing.get('edit_count', 0)
        edit_request = existing.get('edit_request', {})
        
        if edit_count >= 2 and edit_request.get('status') != 'approved':
            return jsonify({'ok': False, 'error': 'Edit limit exceeded. Please request admin approval.'})
            
        if edit_request.get('status') == 'approved':
            data['edit_request'] = None
            
        data['edit_count'] = edit_count + 1
        
        submissions_col.update_one({'_id': ObjectId(sub_id)}, {'$set': data})
        return jsonify({'ok': True, 'id': sub_id})
    else:
        data['edit_count'] = 0
        result = submissions_col.insert_one(data)
        return jsonify({'ok': True, 'id': str(result.inserted_id)})

@app.route('/api/upload-evidence', methods=['POST'])
def upload_evidence():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file part'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'ok': False, 'error': 'No selected file'})
    
    if file:
        filename = secure_filename(file.filename)
        ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        filename = f"{ts}_{filename}"
        
        upload_path = os.path.join('static', 'uploads')
        os.makedirs(upload_path, exist_ok=True)
        
        file.save(os.path.join(upload_path, filename))
        url = url_for('static', filename=f"uploads/{filename}")
        return jsonify({'ok': True, 'url': url})
    
    return jsonify({'ok': False, 'error': 'Unknown error'})

@app.route('/api/request-edit/<sub_id>', methods=['POST'])
def request_edit(sub_id):
    data = request.json
    comment = data.get('comment', '').strip()
    
    submissions_col.update_one(
        {'_id': ObjectId(sub_id)},
        {'$set': {'edit_request': {'pending': True, 'comment': comment, 'status': 'pending', 'timestamp': datetime.utcnow().isoformat()}}}
    )
    return jsonify({'ok': True})

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    
    if not name or not email:
        return jsonify({'ok': False, 'error': 'Name and Email are required'})
        
    existing = users_col.find_one({'email': email})
    if existing:
        return jsonify({'ok': False, 'error': 'You already have an account. Please use Login.'})
        
    users_col.insert_one({'name': name, 'email': email, 'created_at': datetime.utcnow().isoformat(), 'first_time_login': True})
    session['user_email'] = email
    session['user_name'] = name
    return jsonify({'ok': True})

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    email = data.get('email', '').strip()
    
    if not email:
        return jsonify({'ok': False, 'error': 'Email is required'})
        
    user = users_col.find_one({'email': email})
    if not user:
        return jsonify({'ok': False, 'error': 'User not found. Please register as a New User.'})
        
    session['user_email'] = user['email']
    session['user_name'] = user['name']
    return jsonify({'ok': True})

@app.route('/api/change-passcode', methods=['POST'])
def change_passcode():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})
    email = session['user_email']
    user = users_col.find_one({'email': email})
    
    data = request.json
    old_passcode = data.get('old_passcode')
    new_passcode = data.get('new_passcode')
    
    if user.get('passcode') and str(user.get('passcode')) != str(old_passcode):
        return jsonify({'ok': False, 'error': 'Incorrect current passcode'})
        
    if not new_passcode or len(str(new_passcode)) != 4:
        return jsonify({'ok': False, 'error': 'New passcode must be 4 digits'})
        
    users_col.update_one({'email': email}, {'$set': {'passcode': str(new_passcode)}})
    return jsonify({'ok': True})

@app.route('/api/verify-passcode', methods=['POST'])
def verify_passcode():
    email = session.get('pending_email') or session.get('user_email')
    if not email:
        return jsonify({'ok': False, 'error': 'Session expired. Please try logging in again.'})
    passcode = request.json.get('passcode')
    user = users_col.find_one({'email': email})
    
    if not user or str(user.get('passcode')) != str(passcode):
        return jsonify({'ok': False, 'error': 'Invalid passcode'})
        
    session['user_email'] = email
    session['user_name'] = user['name']
    session.pop('pending_email', None)
    return jsonify({'ok': True})

@app.route('/api/forgot-passcode', methods=['POST'])
def forgot_passcode():
    email = request.json.get('email', '').strip()
    user = users_col.find_one({'email': email})
    if not user:
        return jsonify({'ok': False, 'error': 'User not found'})
        
    otp = str(random.randint(1000, 9999))
    users_col.update_one({'email': email}, {'$set': {'reset_otp': otp}})
    
    send_otp_email(email, otp)
    
    session['pending_email'] = email
    return jsonify({'ok': True})

@app.route('/api/reset-passcode', methods=['POST'])
def reset_passcode():
    email = session.get('pending_email') or session.get('user_email')
    if not email:
        return jsonify({'ok': False, 'error': 'Session expired. Please try again.'})
    data = request.json
    otp = data.get('otp')
    new_passcode = data.get('passcode')
    
    if not otp or not new_passcode or len(str(new_passcode)) != 4:
        return jsonify({'ok': False, 'error': 'Invalid data'})
        
    user = users_col.find_one({'email': email})
    if str(user.get('reset_otp')) != str(otp):
        return jsonify({'ok': False, 'error': 'Invalid OTP'})
        
    users_col.update_one({'email': email}, {'$set': {'passcode': str(new_passcode), 'reset_otp': None}})
    session['user_email'] = email
    session['user_name'] = user['name']
    session.pop('pending_email', None)
    return jsonify({'ok': True})

@app.route('/api/update-profile', methods=['POST'])
def update_profile():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})
    dept = request.json.get('department')
    timeout_pref = request.json.get('timeout_pref', 15)
    lock_enabled = request.json.get('lock_enabled', False)
    users_col.update_one({'email': session['user_email']}, {'$set': {
        'department': dept, 
        'timeout_pref': int(timeout_pref), 
        'lock_enabled': bool(lock_enabled),
        'first_time_login': False
    }})
    return jsonify({'ok': True})

@app.route('/api/logout', methods=['POST', 'GET'])
def logout():
    session.pop('user_email', None)
    session.pop('user_name', None)
    if request.method == 'GET':
        return redirect(url_for('login'))
    return jsonify({'ok': True})

@app.route('/api/my-submissions')
def my_submissions():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})
        
    email = session['user_email']
    form_type = request.args.get('type')
    
    query = {'identity.submitterEmail': email}
    if form_type:
        query['form_type'] = form_type
        
    docs = list(submissions_col.find(query, sort=[('timestamp', -1)]))
    for d in docs:
        d['_id'] = str(d['_id'])
    return jsonify({'ok': True, 'submissions': docs})

@app.route('/api/download-template')
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faculty Summary"

    # Styles
    hdr_fill = PatternFill("solid", fgColor="0A2558")
    hdr_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    sub_fill = PatternFill("solid", fgColor="F4A819")
    sub_font = Font(color="1A1A1A", bold=True, size=10, name="Calibri")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:K1')
    ws['A1'] = "JAIN (Deemed-to-be University) — Faculty Readiness Summary"
    ws['A1'].font = Font(color="FFFFFF", bold=True, size=13, name="Calibri")
    ws['A1'].fill = PatternFill("solid", fgColor="0A2558")
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    # Sub-header
    ws.merge_cells('A2:K2')
    ws['A2'] = "Office of Academics — Semester Readiness | Fill one row per faculty-course combination"
    ws['A2'].font = Font(color="1A1A1A", bold=True, size=10, name="Calibri")
    ws['A2'].fill = PatternFill("solid", fgColor="F4A819")
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 20

    headers = [
        "Faculty Name", "Course Name", "Course Code",
        "TLEP Submitted? (Yes/No/Pending)",
        "TLEP Audited by HOD? (Yes/No/Pending)",
        "Session Plan on ERP? (Yes/No/Pending)",
        "LMS Course Live? (Yes/No/Pending)",
        "Program", "Year/Sem",
        "No. of Students",
        "HOD Remarks"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[3].height = 40

    col_widths = [24, 30, 14, 28, 28, 28, 26, 22, 12, 14, 28]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sample rows
    samples = [
        ["Dr. Example Faculty", "Data Structures", "CS301", "Yes", "Yes", "Yes", "Yes", "B.Tech CSE", "3rd / Even", "60", "On track"],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
    ]
    sample_fill = PatternFill("solid", fgColor="F0F4FF")
    for r, row in enumerate(samples, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if r == 4:
                cell.fill = sample_fill
                cell.font = Font(name="Calibri", size=10, italic=True, color="666666")

    # Note
    ws.merge_cells(f'A{4+len(samples)}:K{4+len(samples)}')
    note_cell = ws.cell(row=4+len(samples), column=1, value="📌 Note: Row 4 is a sample. Delete it and fill from Row 5 onwards. Allowed values: Yes / No / Pending")
    note_cell.font = Font(name="Calibri", size=9, italic=True, color="AA4400")
    note_cell.fill = PatternFill("solid", fgColor="FFF8E7")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="Faculty_Summary_Template_OoA.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/api/parse-excel', methods=['POST'])
def parse_excel():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file'})
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'ok': False, 'error': 'Only .xlsx / .xls files accepted'})
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            name = str(row[0] or '').strip()
            if not name or name.startswith('📌') or name.lower() == 'dr. example faculty':
                continue
            rows.append({
                'name': name,
                'course': str(row[1] or '').strip(),
                'code': str(row[2] or '').strip(),
                'tlep': str(row[3] or '').strip(),
                'audited': str(row[4] or '').strip(),
                'erp': str(row[5] or '').strip(),
                'lms': str(row[6] or '').strip(),
                'program': str(row[7] or '').strip(),
                'yearSem': str(row[8] or '').strip(),
                'students': str(row[9] or '').strip(),
                'remarks': str(row[10] or '').strip(),
            })
        return jsonify({'ok': True, 'rows': rows, 'count': len(rows)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/export/<submission_id>')
def export_submission(submission_id):
    try:
        doc = submissions_col.find_one({'_id': ObjectId(submission_id)})
    except:
        return "Not found", 404
    if not doc:
        return "Not found", 404
    wb = build_workbook([doc])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    dept = doc.get('identity', {}).get('dept', 'Department').replace(' ', '_')[:30]
    return send_file(buf, as_attachment=True,
                     download_name=f"SemReadiness_{dept}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Admin ────────────────────────────────────────────────

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        u = request.form.get('username')
        p = request.form.get('password')
        if u == ADMIN_USERNAME and p == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))
        flash('Invalid credentials')
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('admin_login'))

@app.route('/admin')
def admin_dashboard():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    submissions = list(submissions_col.find().sort('timestamp', -1))
    for s in submissions:
        s['_id'] = str(s['_id'])
        
    users = list(users_col.find().sort('created_at', -1))
    for u in users:
        u['_id'] = str(u['_id'])
        latest_sub = submissions_col.find_one({'identity.submitterEmail': u['email']}, sort=[('timestamp', -1)])
        u['latest_dept'] = u.get('department') or (latest_sub['identity']['dept'] if latest_sub and 'identity' in latest_sub else 'N/A')
        
    settings = settings_col.find_one({'_id': 'global'}) or {
        'readiness_enabled': True, 'readiness_deadline': '',
        'closure_enabled': True, 'closure_deadline': ''
    }
    return render_template('admin.html', submissions=submissions, hod_sections=HOD_SECTIONS, users=users, settings=settings, departments=DEPARTMENTS)

@app.route('/admin/settings', methods=['POST'])
def save_settings():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    data = request.json
    settings_col.update_one(
        {'_id': 'global'},
        {'$set': {
            'readiness_enabled': data.get('readiness_enabled', True),
            'readiness_deadline': data.get('readiness_deadline', ''),
            'closure_enabled': data.get('closure_enabled', True),
            'closure_deadline': data.get('closure_deadline', '')
        }},
        upsert=True
    )
    return jsonify({'ok': True})

@app.route('/admin/approve-edit/<sub_id>', methods=['POST'])
def approve_edit(sub_id):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
        
    submissions_col.update_one(
        {'_id': ObjectId(sub_id)},
        {'$set': {'edit_request.status': 'approved', 'edit_request.pending': False}}
    )
    return jsonify({'ok': True})

@app.route('/admin/export-all')
def export_all():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    docs = list(submissions_col.find().sort('timestamp', -1))
    wb = build_workbook(docs)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"SemReadiness_ALL_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/admin/delete/<sid>', methods=['POST'])
def delete_submission(sid):
    if not session.get('admin'):
        return jsonify({'ok': False})
    submissions_col.delete_one({'_id': ObjectId(sid)})
    return jsonify({'ok': True})

def build_workbook(submissions):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill("solid", fgColor="0A2558")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    gold_fill = PatternFill("solid", fgColor="F4A819")
    gold_font = Font(color="1A1A1A", bold=True, size=10, name="Calibri")
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = thin
        ws.row_dimensions[row].height = 36

    def add_title(ws, text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        t = ws.cell(row=1, column=1, value=text)
        t.font = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
        t.fill = hdr_fill
        t.alignment = center
        ws.row_dimensions[1].height = 28

    # Sheet 1 — Dashboard
    ws1 = wb.create_sheet("Dashboard")
    add_title(ws1, "JAIN University — Semester Readiness Dashboard", 22)
    heads = ['Timestamp','Campus','Department','HOD Name','HOD Email','Semester','Acad Year','Sub Date','HOD Deadline',
             'Programs','Courses','Students','Regular Faculty','Adjunct','Labs','Electives','New Courses','On Leave',
             'Fac Checklists','HOD Checklist','HOD Summary','Dashboard']
    style_header(ws1, heads, 2)
    for r, d in enumerate(submissions, 3):
        idt = d.get('identity', {})
        snp = d.get('snapshot', {})
        rs = d.get('readinessStatus', {})
        vals = [d.get('timestamp',''), idt.get('campus',''), idt.get('dept',''), idt.get('hodName',''), idt.get('hodEmail',''),
                idt.get('semester',''), idt.get('acYear',''), idt.get('subDate',''), idt.get('hodDeadline',''),
                snp.get('programs',''), snp.get('courses',''), snp.get('students',''), snp.get('faculty',''),
                snp.get('adjunct',''), snp.get('labs',''), snp.get('electives',''), snp.get('newCourses',''), snp.get('onLeave',''),
                rs.get('faculty',{}).get('submitted',''), rs.get('hod',{}).get('submitted',''),
                rs.get('summary',{}).get('submitted',''), rs.get('dashboard',{}).get('submitted','')]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            cell.fill = PatternFill("solid", fgColor="F5F8FF" if r % 2 == 0 else "FFFFFF")
    ws1.column_dimensions['A'].width = 20
    for i in range(2, 23): ws1.column_dimensions[get_column_letter(i)].width = 18

    # Sheet 2 — Programs
    ws2 = wb.create_sheet("Program Breakdown")
    add_title(ws2, "Program-wise Breakdown", 7)
    style_header(ws2, ['Department','Campus','Program Name','Courses','Students','Faculty','Coordinator'], 2)
    r = 3
    for d in submissions:
        for p in d.get('programs', []):
            vals = [d.get('identity',{}).get('dept',''), d.get('identity',{}).get('campus',''),
                    p.get('name',''), p.get('courses',''), p.get('students',''), p.get('faculty',''), p.get('coord','')]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
            r += 1
    for i, w in enumerate([28,14,32,10,10,10,22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Sheet 3 — Faculty
    ws3 = wb.create_sheet("Faculty Summary")
    add_title(ws3, "Faculty Readiness Summary", 13)
    style_header(ws3, ['Department','Campus','Semester','Faculty Name','Course Name','Code','TLEP Sub?','TLEP Audited?','ERP Plan?','LMS Ready?','Program','Year/Sem','HOD Remarks'], 2)
    r = 3
    for d in submissions:
        for f in d.get('facultySummary', []):
            vals = [d.get('identity',{}).get('dept',''), d.get('identity',{}).get('campus',''), d.get('identity',{}).get('semester',''),
                    f.get('name',''), f.get('course',''), f.get('code',''), f.get('tlep',''), f.get('audited',''),
                    f.get('erp',''), f.get('lms',''), f.get('program',''), f.get('yearSem',''), f.get('remarks','')]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
            r += 1
    for i, w in enumerate([26,14,16,22,28,12,12,14,12,12,24,12,26], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Sheet 4 — HOD Checklist
    ws4 = wb.create_sheet("HOD Checklist")
    add_title(ws4, "HOD Checklist — All Departments", 8)
    style_header(ws4, ['Department','Campus','Semester','Item #','Checklist Item','Section','Status','Remark'], 2)
    r = 3
    for d in submissions:
        chk = d.get('hodChecklist', {})
        for sec in HOD_SECTIONS:
            for item in sec['items']:
                val = chk.get(item['id'], {})
                vals = [d.get('identity',{}).get('dept',''), d.get('identity',{}).get('campus',''),
                        d.get('identity',{}).get('semester',''), item['id'], item['text'],
                        sec['title'].replace('Section ','').split(': ',1)[-1],
                        val.get('status',''), val.get('remark','')]
                for c, v in enumerate(vals, 1):
                    cell = ws4.cell(row=r, column=c, value=v)
                    cell.border = thin
                    cell.font = Font(name="Calibri", size=10)
                    if val.get('status') == 'No':
                        cell.fill = PatternFill("solid", fgColor="FFE8E8")
                    elif val.get('status') == 'Yes':
                        cell.fill = PatternFill("solid", fgColor="E8F8EE")
                r += 1
    for i, w in enumerate([26,12,14,8,52,24,10,30], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # Sheet 5 — Gap Analysis
    ws5 = wb.create_sheet("Gap Analysis")
    add_title(ws5, "Gap Analysis — Checklist Completion by Department", 9)
    style_header(ws5, ['Department','Campus','Semester','Yes','No','N/A','Blank','Completion %','Gap Items (No)'], 2)
    all_items = [item for sec in HOD_SECTIONS for item in sec['items']]
    r = 3
    for d in submissions:
        chk = d.get('hodChecklist', {})
        vals_list = list(chk.values())
        yes = sum(1 for x in vals_list if x.get('status') == 'Yes')
        no = sum(1 for x in vals_list if x.get('status') == 'No')
        na = sum(1 for x in vals_list if x.get('status') == 'N/A')
        blank = sum(1 for x in vals_list if not x.get('status'))
        denom = (yes + no + blank)
        pct = f"{round(yes/denom*100)}%" if denom else "—"
        gaps = "; ".join(f"{i['id']}. {i['text']}" for i in all_items if chk.get(i['id'], {}).get('status') == 'No')
        row_vals = [d.get('identity',{}).get('dept',''), d.get('identity',{}).get('campus',''),
                    d.get('identity',{}).get('semester',''), yes, no, na, blank, pct, gaps]
        for c, v in enumerate(row_vals, 1):
            cell = ws5.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if c == 5 and v and v > 0:
                cell.fill = PatternFill("solid", fgColor="FFE8E8")
                cell.font = Font(name="Calibri", size=10, bold=True, color="AA0000")
        r += 1
    for i, w in enumerate([26,12,14,8,8,8,8,14,60], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    return wb

if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)
    app.run(debug=True, port=5000)
